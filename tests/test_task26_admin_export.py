"""
Test for Task 26: 修改管理员后台导出功能
验证管理员后台导出包含记录类型列
"""

import os
import tempfile
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from src.models import Invoice
from src.export_service import ExportService


def test_admin_export_includes_record_type_column():
    """测试管理员后台导出包含记录类型列"""
    # 创建混合类型的测试数据（模拟管理员后台的数据）
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='购买文具',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='张三',
            reimbursement_person_id=1,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-20251228-100000-A1B2',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='打车费用',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='李四',
            reimbursement_person_id=2,
            reimbursement_status='未报销',
            record_type='manual'
        ),
        Invoice(
            invoice_number='INV-002',
            invoice_date='2025-12-29',
            item_name='餐饮费',
            amount=Decimal('200.00'),
            remark='团队聚餐',
            file_path='/path/to/pdf2',
            scan_time=datetime(2025, 12, 29, 10, 0, 0),
            uploaded_by='王五',
            reimbursement_person_id=1,
            reimbursement_status='已报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-20251229-120000-C3D4',
            invoice_date='2025-12-29',
            item_name='住宿费',
            amount=Decimal('300.00'),
            remark='出差住宿',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 29, 12, 0, 0),
            uploaded_by='赵六',
            reimbursement_person_id=3,
            reimbursement_status='已报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证表头包含"记录类型"列 (Requirement 13.7)
        headers = [cell.value for cell in ws[1]]
        assert '记录类型' in headers, "管理员后台导出应包含'记录类型'列"
        
        # 获取记录类型列的索引
        record_type_col_idx = headers.index('记录类型') + 1
        
        # 验证第一条记录（发票）显示为"发票" (Requirement 13.7)
        assert ws.cell(row=2, column=record_type_col_idx).value == '发票', \
            "发票记录应显示为'发票'"
        
        # 验证第二条记录（手动记录）显示为"无票报销" (Requirement 13.7)
        assert ws.cell(row=3, column=record_type_col_idx).value == '无票报销', \
            "手动记录应显示为'无票报销'"
        
        # 验证第三条记录（发票）
        assert ws.cell(row=4, column=record_type_col_idx).value == '发票'
        
        # 验证第四条记录（手动记录）
        assert ws.cell(row=5, column=record_type_col_idx).value == '无票报销'
        
        print("✓ 管理员后台导出包含记录类型列，显示为'发票'或'无票报销'")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_admin_export_all_records():
    """测试管理员后台导出包含所有记录"""
    # 创建测试数据
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='张三',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-001',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='李四',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证所有记录都被导出 (Requirement 13.7)
        assert ws.max_row >= 3, "应该导出所有记录（2条数据 + 1行表头）"
        
        # 验证发票号码列包含所有记录
        invoice_numbers = [ws.cell(row=i, column=1).value for i in range(2, 4)]
        assert 'INV-001' in invoice_numbers
        assert 'MANUAL-001' in invoice_numbers
        
        print("✓ 管理员后台导出包含所有记录")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_admin_export_statistics():
    """测试管理员后台导出包含分类统计"""
    # 创建混合类型的测试数据
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='张三',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='INV-002',
            invoice_date='2025-12-28',
            item_name='餐饮费',
            amount=Decimal('200.00'),
            remark='',
            file_path='/path/to/pdf2',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='李四',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-001',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 12, 0, 0),
            uploaded_by='王五',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 查找统计信息行
        summary_row = 3 + 3  # 3条数据 + 1行表头 + 2行空白
        
        # 验证总计统计
        assert '汇总统计' in str(ws.cell(row=summary_row, column=1).value)
        assert '总记录数: 3' in str(ws.cell(row=summary_row, column=2).value)
        
        # 验证发票记录统计（显示"发票记录"）
        assert '发票记录: 2张' in str(ws.cell(row=summary_row + 1, column=2).value)
        
        # 验证手动记录统计（显示"无票报销记录"）
        assert '无票报销记录: 1张' in str(ws.cell(row=summary_row + 2, column=2).value)
        
        print("✓ 管理员后台导出包含分类统计")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


if __name__ == "__main__":
    print("运行Task 26测试...\n")
    test_admin_export_includes_record_type_column()
    test_admin_export_all_records()
    test_admin_export_statistics()
    print("\n所有测试通过！✓")
