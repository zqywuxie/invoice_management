"""
Test for Task 9: 修改导出功能支持手动记录
"""

import os
import tempfile
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from src.models import Invoice
from src.export_service import ExportService


def test_export_includes_record_type_column():
    """测试导出包含记录类型列"""
    # 创建测试数据
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='购买文具',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-20251228-100000-A1B2',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='打车费用',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证表头包含"记录类型"列
        headers = [cell.value for cell in ws[1]]
        assert '记录类型' in headers, "导出文件应包含'记录类型'列"
        
        # 获取记录类型列的索引
        record_type_col_idx = headers.index('记录类型') + 1
        
        # 验证第一条记录（发票）
        assert ws.cell(row=2, column=record_type_col_idx).value == '发票'
        
        # 验证第二条记录（手动记录）
        assert ws.cell(row=3, column=record_type_col_idx).value == '无票报销'
        
        print("✓ 导出包含记录类型列")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_includes_all_record_types():
    """测试导出包含所有记录类型"""
    # 创建混合类型的测试数据
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-001',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        ),
        Invoice(
            invoice_number='INV-002',
            invoice_date='2025-12-29',
            item_name='餐饮费',
            amount=Decimal('200.00'),
            remark='',
            file_path='/path/to/pdf2',
            scan_time=datetime(2025, 12, 29, 10, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-002',
            invoice_date='2025-12-29',
            item_name='住宿费',
            amount=Decimal('300.00'),
            remark='',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 29, 11, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证所有记录都被导出（4条数据 + 1行表头）
        assert ws.max_row >= 5, "应该导出所有记录"
        
        # 验证发票号码列包含所有记录的ID
        invoice_numbers = [ws.cell(row=i, column=1).value for i in range(2, 6)]
        assert 'INV-001' in invoice_numbers
        assert 'MANUAL-001' in invoice_numbers
        assert 'INV-002' in invoice_numbers
        assert 'MANUAL-002' in invoice_numbers
        
        print("✓ 导出包含所有记录类型")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_manual_record_uses_generated_identifier():
    """测试导出手动记录时使用生成的标识符"""
    # 创建手动记录
    manual_record = Invoice(
        invoice_number='MANUAL-20251228-143052-A3F2',
        invoice_date='2025-12-28',
        item_name='交通费',
        amount=Decimal('50.00'),
        remark='打车费用',
        file_path='MANUAL',
        scan_time=datetime(2025, 12, 28, 14, 30, 52),
        uploaded_by='测试用户',
        reimbursement_person_id=None,
        reimbursement_status='未报销',
        record_type='manual'
    )
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel([manual_record], output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证发票号码列包含生成的标识符
        invoice_number = ws.cell(row=2, column=1).value
        assert invoice_number == 'MANUAL-20251228-143052-A3F2'
        assert invoice_number.startswith('MANUAL-')
        
        print("✓ 导出手动记录时使用生成的标识符")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_includes_all_fields():
    """测试导出包含所有字段"""
    # 创建测试数据
    invoice = Invoice(
        invoice_number='INV-001',
        invoice_date='2025-12-28',
        item_name='办公用品',
        amount=Decimal('100.50'),
        remark='购买文具',
        file_path='/path/to/pdf',
        scan_time=datetime(2025, 12, 28, 10, 30, 45),
        uploaded_by='测试用户',
        reimbursement_person_id=None,
        reimbursement_status='未报销',
        record_type='invoice'
    )
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel([invoice], output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证所有字段都存在
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['发票号码', '记录类型', '开票日期', '项目名称', '金额', '备注', '源文件路径', '扫描时间']
        
        for expected_header in expected_headers:
            assert expected_header in headers, f"应包含'{expected_header}'列"
        
        # 验证数据行包含所有字段值
        assert ws.cell(row=2, column=1).value == 'INV-001'
        assert ws.cell(row=2, column=2).value == '发票'
        assert ws.cell(row=2, column=3).value == '2025-12-28'
        assert ws.cell(row=2, column=4).value == '办公用品'
        assert ws.cell(row=2, column=5).value == '100.50'
        assert ws.cell(row=2, column=6).value == '购买文具'
        assert ws.cell(row=2, column=7).value == '/path/to/pdf'
        assert ws.cell(row=2, column=8).value == '2025-12-28 10:30:45'
        
        print("✓ 导出包含所有字段")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_statistics_by_record_type():
    """测试导出包含按记录类型分类的统计信息"""
    # 创建混合类型的测试数据
    invoices = [
        Invoice(
            invoice_number='INV-001',
            invoice_date='2025-12-28',
            item_name='办公用品',
            amount=Decimal('100.00'),
            remark='',
            file_path='/path/to/pdf',
            scan_time=datetime(2025, 12, 28, 10, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='INV-002',
            invoice_date='2025-12-28',
            item_name='餐饮费',
            amount=Decimal('200.00'),
            remark='',
            file_path='/path/to/pdf2',
            scan_time=datetime(2025, 12, 28, 11, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='invoice'
        ),
        Invoice(
            invoice_number='MANUAL-001',
            invoice_date='2025-12-28',
            item_name='交通费',
            amount=Decimal('50.00'),
            remark='',
            file_path='MANUAL',
            scan_time=datetime(2025, 12, 28, 12, 0, 0),
            uploaded_by='测试用户',
            reimbursement_person_id=None,
            reimbursement_status='未报销',
            record_type='manual'
        )
    ]
    
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel(invoices, output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 查找统计信息行（数据行后面 + 2行空白）
        summary_row = 3 + 3  # 3条数据 + 1行表头 + 2行空白
        
        # 验证总计统计
        assert '汇总统计' in str(ws.cell(row=summary_row, column=1).value)
        assert '总记录数: 3' in str(ws.cell(row=summary_row, column=2).value)
        assert '350.00' in str(ws.cell(row=summary_row, column=4).value)
        
        # 验证发票记录统计
        assert '发票记录: 2张' in str(ws.cell(row=summary_row + 1, column=2).value)
        assert '300.00' in str(ws.cell(row=summary_row + 1, column=4).value)
        
        # 验证手动记录统计
        assert '无票报销记录: 1张' in str(ws.cell(row=summary_row + 2, column=2).value)
        assert '50.00' in str(ws.cell(row=summary_row + 2, column=4).value)
        
        print("✓ 导出包含按记录类型分类的统计信息")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_empty_list():
    """测试导出空列表"""
    # 导出到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    try:
        export_service = ExportService()
        export_service.export_to_excel([], output_path)
        
        # 读取导出的文件
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 验证只有表头行
        assert ws.max_row >= 1
        
        # 验证表头存在
        headers = [cell.value for cell in ws[1]]
        assert '记录类型' in headers
        
        print("✓ 导出空列表成功")
        
    finally:
        # 清理临时文件
        if os.path.exists(output_path):
            os.remove(output_path)


if __name__ == "__main__":
    print("运行Task 9实现测试...\n")
    test_export_includes_record_type_column()
    test_export_includes_all_record_types()
    test_export_manual_record_uses_generated_identifier()
    test_export_includes_all_fields()
    test_export_statistics_by_record_type()
    test_export_empty_list()
    print("\n所有测试通过！✓")
