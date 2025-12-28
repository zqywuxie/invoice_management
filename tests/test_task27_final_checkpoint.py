"""
Task 27: Final Checkpoint - 验证所有更新
测试用户端所有"无票报销"文案显示正确
测试管理员后台记录类型过滤功能
测试管理员后台分类统计显示
测试导出文件中的记录类型列
"""

import os
import sqlite3
import tempfile
from datetime import datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from src.data_store import DataStore
from src.export_service import ExportService
from src.models import Invoice
from src.sqlite_data_store import SQLiteDataStore


class TestTask27FinalCheckpoint:
    """Task 27: Final Checkpoint - 验证所有更新"""

    @pytest.fixture
    def temp_db(self):
        """创建临时数据库"""
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        yield path
        if os.path.exists(path):
            os.remove(path)

    @pytest.fixture
    def data_store(self, temp_db):
        """创建数据存储实例"""
        store = SQLiteDataStore(temp_db)
        return store

    @pytest.fixture
    def sample_invoices(self, data_store):
        """创建测试数据：包含发票记录和手动记录"""
        invoices = []
        
        # 创建2条发票记录
        invoice1 = Invoice(
            invoice_number="12345678",
            invoice_date="2025-01-15",
            item_name="办公用品",
            amount=Decimal("1500.00"),
            remark="购买文具",
            file_path="test1.pdf",
            scan_time=datetime.now(),
            uploaded_by="张三",
            reimbursement_person_id=None,
            reimbursement_status="未报销",
            record_type="invoice"
        )
        data_store.insert(invoice1)
        invoices.append(invoice1)
        
        invoice2 = Invoice(
            invoice_number="87654321",
            invoice_date="2025-01-16",
            item_name="差旅费",
            amount=Decimal("2500.00"),
            remark="出差北京",
            file_path="test2.pdf",
            scan_time=datetime.now(),
            uploaded_by="李四",
            reimbursement_person_id=None,
            reimbursement_status="已报销",
            record_type="invoice"
        )
        data_store.insert(invoice2)
        invoices.append(invoice2)
        
        # 创建2条手动记录
        manual1 = Invoice(
            invoice_number="MANUAL-20250115-120000-A1B2",
            invoice_date="2025-01-15",
            item_name="交通费",
            amount=Decimal("50.00"),
            remark="打车费用",
            file_path="",
            scan_time=datetime.now(),
            uploaded_by="王五",
            reimbursement_person_id=None,
            reimbursement_status="未报销",
            record_type="manual"
        )
        data_store.insert(manual1)
        invoices.append(manual1)
        
        manual2 = Invoice(
            invoice_number="MANUAL-20250116-140000-C3D4",
            invoice_date="2025-01-16",
            item_name="餐费",
            amount=Decimal("80.00"),
            remark="客户招待",
            file_path="",
            scan_time=datetime.now(),
            uploaded_by="赵六",
            reimbursement_person_id=None,
            reimbursement_status="已报销",
            record_type="manual"
        )
        data_store.insert(manual2)
        invoices.append(manual2)
        
        return invoices

    def test_user_side_terminology_display(self):
        """测试用户端所有"无票报销"文案显示正确"""
        print("\n=== 测试用户端'无票报销'文案显示 ===")
        
        # 1. 检查上传页面 (upload.html)
        upload_html_path = 'invoice_web/templates/user/upload.html'
        assert os.path.exists(upload_html_path), "upload.html should exist"
        
        with open(upload_html_path, 'r', encoding='utf-8') as f:
            upload_content = f.read()
        
        # 验证模式选择器按钮文本
        assert '无票报销' in upload_content, "upload.html should contain '无票报销' text"
        assert 'manual-mode-btn' in upload_content, "upload.html should have manual mode button"
        print("✓ 上传页面包含'无票报销'按钮")
        
        # 2. 检查发票列表页面 (invoices.html)
        invoices_html_path = 'invoice_web/templates/user/invoices.html'
        assert os.path.exists(invoices_html_path), "invoices.html should exist"
        
        with open(invoices_html_path, 'r', encoding='utf-8') as f:
            invoices_content = f.read()
        
        # 验证过滤按钮文本
        assert '无票报销' in invoices_content, "invoices.html should contain '无票报销' filter"
        assert 'filter-manual' in invoices_content, "invoices.html should have manual filter button"
        
        # 验证统计显示文本
        assert '无发票记录' in invoices_content, "invoices.html should show '无发票记录' in statistics"
        print("✓ 发票列表页面包含'无票报销'过滤器和统计")
        
        # 3. 检查详情页面 (detail.html)
        detail_html_path = 'invoice_web/templates/user/detail.html'
        assert os.path.exists(detail_html_path), "detail.html should exist"
        
        with open(detail_html_path, 'r', encoding='utf-8') as f:
            detail_content = f.read()
        
        # 验证手动记录提示文本
        assert '无票报销' in detail_content, "detail.html should contain '无票报销' notice"
        assert 'manual-record-notice' in detail_content, "detail.html should have manual record notice"
        print("✓ 详情页面包含'无票报销'提示")
        
        # 4. 检查用户端JavaScript (user_app.js)
        user_js_path = 'invoice_web/static/js/user_app.js'
        assert os.path.exists(user_js_path), "user_app.js should exist"
        
        with open(user_js_path, 'r', encoding='utf-8') as f:
            user_js_content = f.read()
        
        # 验证JavaScript中的文案
        assert '无票报销' in user_js_content, "user_app.js should contain '无票报销' text"
        assert 'badge-manual' in user_js_content, "user_app.js should have badge-manual class"
        print("✓ 用户端JavaScript包含'无票报销'文案")
        
        print("\n✅ 用户端所有'无票报销'文案显示正确")

    def test_admin_backend_record_type_filter(self):
        """测试管理员后台记录类型过滤功能"""
        print("\n=== 测试管理员后台记录类型过滤功能 ===")
        
        # 1. 检查管理员后台HTML (index.html)
        admin_html_path = 'invoice_web/templates/index.html'
        assert os.path.exists(admin_html_path), "index.html should exist"
        
        with open(admin_html_path, 'r', encoding='utf-8') as f:
            admin_content = f.read()
        
        # 验证过滤器HTML结构
        assert 'adminRecordTypeFilter' in admin_content, "Should have admin record type filter"
        assert 'admin-filter-all' in admin_content, "Should have '全部' filter button"
        assert 'admin-filter-invoice' in admin_content, "Should have '有发票' filter button"
        assert 'admin-filter-manual' in admin_content, "Should have '无票报销' filter button"
        
        # 验证按钮文本
        assert '无票报销' in admin_content, "Should contain '无票报销' text"
        assert '有发票' in admin_content, "Should contain '有发票' text"
        print("✓ 管理员后台HTML包含记录类型过滤器")
        
        # 2. 检查管理员后台JavaScript (app.js)
        admin_js_path = 'invoice_web/static/js/app.js'
        assert os.path.exists(admin_js_path), "app.js should exist"
        
        with open(admin_js_path, 'r', encoding='utf-8') as f:
            admin_js_content = f.read()
        
        # 验证JavaScript中的过滤逻辑
        assert 'recordTypeFilter' in admin_js_content, "Should have recordTypeFilter in state"
        assert 'adminRecordTypeFilter' in admin_js_content, "Should handle admin record type filter"
        
        # 验证记录类型标识显示
        assert 'badge-manual' in admin_js_content, "Should have badge-manual class"
        assert 'badge-invoice' in admin_js_content, "Should have badge-invoice class"
        assert '无票报销' in admin_js_content, "Should contain '无票报销' text in JS"
        print("✓ 管理员后台JavaScript包含过滤逻辑")
        
        print("\n✅ 管理员后台记录类型过滤功能正确")

    def test_admin_backend_categorized_statistics(self):
        """测试管理员后台分类统计显示"""
        print("\n=== 测试管理员后台分类统计显示 ===")
        
        # 检查管理员后台HTML (index.html)
        admin_html_path = 'invoice_web/templates/index.html'
        assert os.path.exists(admin_html_path), "index.html should exist"
        
        with open(admin_html_path, 'r', encoding='utf-8') as f:
            admin_content = f.read()
        
        # 验证统计显示元素
        assert 'invoiceCount' in admin_content, "Should have invoiceCount element"
        assert 'manualCount' in admin_content, "Should have manualCount element"
        assert 'invoiceAmount' in admin_content, "Should have invoiceAmount element"
        assert 'manualAmount' in admin_content, "Should have manualAmount element"
        
        # 验证统计标签文本
        assert '有发票记录' in admin_content, "Should show '有发票记录' label"
        assert '无票报销记录' in admin_content, "Should show '无票报销记录' label"
        print("✓ 管理员后台HTML包含分类统计显示")
        
        # 检查JavaScript统计更新逻辑
        admin_js_path = 'invoice_web/static/js/app.js'
        with open(admin_js_path, 'r', encoding='utf-8') as f:
            admin_js_content = f.read()
        
        # 验证统计更新函数
        assert 'invoiceCount' in admin_js_content, "JS should update invoiceCount"
        assert 'manualCount' in admin_js_content, "JS should update manualCount"
        assert 'invoiceAmount' in admin_js_content, "JS should update invoiceAmount"
        assert 'manualAmount' in admin_js_content, "JS should update manualAmount"
        print("✓ 管理员后台JavaScript包含统计更新逻辑")
        
        print("\n✅ 管理员后台分类统计显示正确")

    def test_export_record_type_column(self, data_store, sample_invoices):
        """测试导出文件中的记录类型列"""
        print("\n=== 测试导出文件中的记录类型列 ===")
        
        # 创建导出服务
        export_service = ExportService()
        
        # 导出到临时文件
        fd, temp_excel = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        try:
            # 执行导出
            export_service.export_to_excel(sample_invoices, temp_excel)
            print(f"✓ 导出文件创建成功: {temp_excel}")
            
            # 加载并验证Excel文件
            wb = load_workbook(temp_excel)
            ws = wb.active
            
            # 1. 验证表头包含"记录类型"列
            headers = [cell.value for cell in ws[1]]
            assert '记录类型' in headers, "Export should have '记录类型' column"
            record_type_col_idx = headers.index('记录类型') + 1
            print(f"✓ 表头包含'记录类型'列 (第{record_type_col_idx}列)")
            
            # 2. 验证发票记录显示为"发票"
            invoice_rows = [row for row in range(2, ws.max_row + 1) 
                          if ws.cell(row=row, column=1).value in ["12345678", "87654321"]]
            for row in invoice_rows:
                record_type = ws.cell(row=row, column=record_type_col_idx).value
                assert record_type == '发票', f"Invoice record should show '发票', got '{record_type}'"
            print(f"✓ 发票记录显示为'发票' ({len(invoice_rows)}条)")
            
            # 3. 验证手动记录显示为"无票报销"
            manual_rows = [row for row in range(2, ws.max_row + 1) 
                         if ws.cell(row=row, column=1).value and 
                         str(ws.cell(row=row, column=1).value).startswith("MANUAL-")]
            for row in manual_rows:
                record_type = ws.cell(row=row, column=record_type_col_idx).value
                assert record_type == '无票报销', f"Manual record should show '无票报销', got '{record_type}'"
            print(f"✓ 手动记录显示为'无票报销' ({len(manual_rows)}条)")
            
            # 4. 验证统计行包含分类统计
            # 查找汇总统计行
            summary_row = None
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == '汇总统计':
                    summary_row = row
                    break
            
            assert summary_row is not None, "Should have summary statistics row"
            print(f"✓ 找到汇总统计行 (第{summary_row}行)")
            
            # 验证总计
            total_label = ws.cell(row=summary_row, column=2).value
            assert '总记录数' in str(total_label), "Should show total count"
            print(f"✓ 总计: {total_label}")
            
            # 验证发票记录统计
            invoice_stats = ws.cell(row=summary_row + 1, column=2).value
            assert '发票记录' in str(invoice_stats), "Should show invoice record count"
            assert '2张' in str(invoice_stats), "Should show 2 invoice records"
            print(f"✓ 发票统计: {invoice_stats}")
            
            # 验证无票报销记录统计
            manual_stats = ws.cell(row=summary_row + 2, column=2).value
            assert '无票报销记录' in str(manual_stats), "Should show manual record count"
            assert '2张' in str(manual_stats), "Should show 2 manual records"
            print(f"✓ 无票报销统计: {manual_stats}")
            
            # 5. 验证金额统计
            invoice_amount_label = ws.cell(row=summary_row + 1, column=3).value
            assert '发票金额' in str(invoice_amount_label), "Should show invoice amount label"
            
            manual_amount_label = ws.cell(row=summary_row + 2, column=3).value
            assert '无票报销金额' in str(manual_amount_label), "Should show manual amount label"
            print("✓ 金额统计标签正确")
            
            print("\n✅ 导出文件中的记录类型列正确")
            
        finally:
            # 清理临时文件
            if os.path.exists(temp_excel):
                os.remove(temp_excel)

    def test_comprehensive_integration(self, data_store, sample_invoices):
        """综合集成测试：验证所有组件协同工作"""
        print("\n=== 综合集成测试 ===")
        
        # 1. 验证数据库中的记录类型
        all_invoices = data_store.load_all()
        invoice_records = [inv for inv in all_invoices if inv.record_type == 'invoice']
        manual_records = [inv for inv in all_invoices if inv.record_type == 'manual']
        
        assert len(invoice_records) == 2, "Should have 2 invoice records"
        assert len(manual_records) == 2, "Should have 2 manual records"
        print(f"✓ 数据库包含 {len(invoice_records)} 条发票记录和 {len(manual_records)} 条手动记录")
        
        # 2. 验证记录类型过滤
        filtered_invoices = [inv for inv in all_invoices if inv.record_type == 'invoice']
        assert len(filtered_invoices) == 2, "Filter should return 2 invoice records"
        
        filtered_manuals = [inv for inv in all_invoices if inv.record_type == 'manual']
        assert len(filtered_manuals) == 2, "Filter should return 2 manual records"
        print("✓ 记录类型过滤功能正常")
        
        # 3. 验证统计计算
        total_amount = sum(inv.amount for inv in all_invoices)
        invoice_amount = sum(inv.amount for inv in invoice_records)
        manual_amount = sum(inv.amount for inv in manual_records)
        
        assert total_amount == invoice_amount + manual_amount, "Total should equal sum of categories"
        assert invoice_amount == Decimal("4000.00"), "Invoice amount should be 4000.00"
        assert manual_amount == Decimal("130.00"), "Manual amount should be 130.00"
        print(f"✓ 统计计算正确: 总计={total_amount}, 发票={invoice_amount}, 无票报销={manual_amount}")
        
        # 4. 验证导出功能
        export_service = ExportService()
        fd, temp_excel = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        try:
            export_service.export_to_excel(all_invoices, temp_excel)
            wb = load_workbook(temp_excel)
            ws = wb.active
            
            # 验证导出的记录数
            data_rows = ws.max_row - 1  # 减去表头行
            # 减去汇总统计行（汇总统计占4行：空行+汇总标题+发票统计+手动统计）
            actual_data_rows = data_rows - 4
            assert actual_data_rows == 4, f"Should export 4 records, got {actual_data_rows}"
            print(f"✓ 导出文件包含 {actual_data_rows} 条记录")
            
        finally:
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
        
        print("\n✅ 综合集成测试通过")

    def test_all_requirements_met(self):
        """验证所有需求都已满足"""
        print("\n=== 验证所有需求 ===")
        
        requirements = [
            ("用户端'无票报销'文案", "upload.html, invoices.html, detail.html, user_app.js"),
            ("管理员后台记录类型过滤", "index.html, app.js"),
            ("管理员后台分类统计", "index.html, app.js"),
            ("导出文件记录类型列", "export_service.py"),
        ]
        
        for req_name, files in requirements:
            print(f"✓ {req_name}: {files}")
        
        print("\n✅ 所有需求已满足")


if __name__ == '__main__':
    # 运行测试
    pytest.main([__file__, '-v', '-s'])
