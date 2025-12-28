"""
ExportService for exporting invoice data to Excel.
导出服务 - 负责将发票数据导出到Excel文件
"""

from decimal import Decimal
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import Invoice


class ExportService:
    """
    导出服务类，负责将发票数据导出到Excel文件
    """
    
    def format_amount(self, amount: Decimal) -> str:
        """
        格式化金额为两位小数的字符串
        
        Args:
            amount: Decimal金额值
            
        Returns:
            格式化后的字符串，包含恰好两位小数
        """
        return f"{amount:.2f}"
    
    def export_to_excel(self, invoices: List[Invoice], output_path: str) -> None:
        """
        导出发票数据到Excel文件
        
        Args:
            invoices: 发票列表
            output_path: 输出文件路径
            
        Raises:
            IOError: 文件写入失败时抛出
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "发票汇总"
        
        # Define headers
        headers = [
            "发票号码",
            "记录类型",
            "开票日期", 
            "项目名称",
            "金额",
            "备注",
            "源文件路径",
            "扫描时间"
        ]

        # Style for headers
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write invoice data
        for row, invoice in enumerate(invoices, 2):
            # 发票号码
            ws.cell(row=row, column=1, value=invoice.invoice_number).border = thin_border
            
            # 记录类型：显示"发票"或"无票报销"
            record_type_display = "无票报销" if invoice.record_type == "manual" else "发票"
            ws.cell(row=row, column=2, value=record_type_display).border = thin_border
            
            # 其他字段
            ws.cell(row=row, column=3, value=invoice.invoice_date).border = thin_border
            ws.cell(row=row, column=4, value=invoice.item_name).border = thin_border
            ws.cell(row=row, column=5, value=self.format_amount(invoice.amount)).border = thin_border
            ws.cell(row=row, column=6, value=invoice.remark).border = thin_border
            ws.cell(row=row, column=7, value=invoice.file_path).border = thin_border
            ws.cell(row=row, column=8, value=invoice.scan_time.strftime("%Y-%m-%d %H:%M:%S")).border = thin_border
        
        # Add summary statistics row
        summary_row = len(invoices) + 3
        
        # Calculate statistics by record type
        total_amount = Decimal("0")
        invoice_count = 0
        manual_count = 0
        invoice_amount = Decimal("0")
        manual_amount = Decimal("0")
        
        for inv in invoices:
            total_amount += inv.amount
            if inv.record_type == 'manual':
                manual_count += 1
                manual_amount += inv.amount
            else:
                invoice_count += 1
                invoice_amount += inv.amount
        
        # Summary label
        ws.cell(row=summary_row, column=1, value="汇总统计").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"总记录数: {len(invoices)}")
        ws.cell(row=summary_row, column=3, value="总金额:")
        ws.cell(row=summary_row, column=4, value=self.format_amount(total_amount)).font = Font(bold=True)
        
        # Detailed statistics
        ws.cell(row=summary_row + 1, column=1, value="")
        ws.cell(row=summary_row + 1, column=2, value=f"发票记录: {invoice_count}张")
        ws.cell(row=summary_row + 1, column=3, value="发票金额:")
        ws.cell(row=summary_row + 1, column=4, value=self.format_amount(invoice_amount))
        
        ws.cell(row=summary_row + 2, column=1, value="")
        ws.cell(row=summary_row + 2, column=2, value=f"无票报销记录: {manual_count}张")
        ws.cell(row=summary_row + 2, column=3, value="无票报销金额:")
        ws.cell(row=summary_row + 2, column=4, value=self.format_amount(manual_amount))
        
        # Adjust column widths
        column_widths = [20, 12, 15, 30, 15, 30, 40, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            raise IOError(f"Failed to export to Excel file {output_path}: {e}")
